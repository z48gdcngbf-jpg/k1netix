"""
Layer 4 — Pre-RFI Generator & Router
Synthesises Layer 3 compliance results into structured Pre-RFI documents.
Routes to responsible professionals via email and stores records.
"""

from __future__ import annotations
import io
import json
import os
import smtplib
import ssl
from collections import Counter
from datetime import datetime, timedelta
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.utils import formataddr
from pathlib import Path
from typing import Any, Optional

import pandas as pd
import streamlit as st

try:
    from deepseek_client import DeepSeekClient
except Exception:  # Layer 4 can still run with deterministic fallback text.
    DeepSeekClient = None

# certifi is bundled with pandas/requests — fallback for SSL on macOS
try:
    import certifi
    _CERT_PATH = certifi.where()
except ImportError:
    _CERT_PATH = None


# ── Discipline config ─────────────────────────────────────────────────────────

DISCIPLINES = {
    "MECH":  {"label": "Mechanical",      "color": "#4fc3f7"},
    "ELEC":  {"label": "Electrical",      "color": "#fff176"},
    "PLMB":  {"label": "Plumbing",        "color": "#80cbc4"},
    "STR":   {"label": "Structural",      "color": "#ef9a9a"},
    "ARCH":  {"label": "Architecture",    "color": "#ce93d8"},
    "CIVIL": {"label": "Civil",           "color": "#a5d6a7"},
    "FP":    {"label": "Fire Protection", "color": "#ffab91"},
    "MEP":   {"label": "MEP",             "color": "#4fc3f7"},
}

DISC_EMAIL_MAP = {
    "MECH": "MECH", "ELEC": "ELEC", "PLMB": "PLMB",
    "MEP":  "MECH", "STR":  "STR",  "ARCH": "ARCH",
    "CIVIL":"CIVIL","FP":   "FP",
}


# ── Pre-RFI sections (expanded to match the user's template) ──────────────────

PRE_RFI_SECTIONS = [
    ("1. PROJECT INFORMATION", [
        "project_name", "project_location", "discipline_label",
        "project_id", "bim_model_version",
        "date_generated", "rfi_number",
    ]),
    ("2. TRACEABILITY & CLASH IDENTIFICATION", [
        "traceability_id", "bcf_topic_guid", "clash_source_tool",
        "floor_zone", "element_a_guid", "element_a_type",
        "element_b_guid", "element_b_type",
        "clash_coordinates", "cluster_size",
        "cluster_label", "clash_pair",
    ]),
    ("3. REGULATION & COMPLIANCE REFERENCE", [
        "applicable_codes", "clause_references", "regulation_version",
        "regulation_quote", "life_safety_related",
        "regulation_match_quality",
    ]),
    ("4. AI REASONING CHAIN", [
        "clash_description", "issue_titles",
        "primary_concern", "compliance_status", "llm_certainty",
        "constructability_assessment", "compliance_summary",
        "priority_band", "priority_score",
        "noise_filter_result", "discipline_confidence",
    ]),
    ("5. RECOMMENDED ACTIONS", [
        "recommended_action", "suggested_solutions",
        "requires_specialist", "specialist_discipline",
    ]),
    ("6. ROUTING & SIGN-OFF", [
        "assigned_to", "assigned_email", "response_required_by",
        "prepared_by", "status", "data_gaps", "notes",
    ]),
]

FIELD_LABELS = {
    "project_name":             "Project Name",
    "project_location":         "Project Location",
    "discipline_label":         "Discipline(s)",
    "project_id":               "Project ID",
    "bim_model_version":        "BIM Model Version",
    "date_generated":           "Date Generated",
    "rfi_number":               "Pre-RFI Number",
    "traceability_id":          "Traceability ID",
    "bcf_topic_guid":           "BCF Topic GUID",
    "clash_source_tool":        "Clash Source Tool",
    "floor_zone":               "Floor / Zone",
    "element_a_guid":           "Element A (GUID)",
    "element_a_type":           "Element A Type",
    "element_b_guid":           "Element B (GUID)",
    "element_b_type":           "Element B Type",
    "clash_coordinates":        "Clash Coordinates (X, Y, Z)",
    "cluster_size":             "Cluster Size",
    "cluster_label":            "Clash Group Label",
    "clash_pair":               "Clash Pair (Disciplines)",
    "applicable_codes":         "Applicable Code / Standard",
    "clause_references":        "Clause / Section Number",
    "regulation_version":       "Version / Year",
    "regulation_quote":         "Regulation Quote",
    "life_safety_related":      "Life-Safety Related?",
    "regulation_match_quality": "Regulation Match Quality",
    "clash_description":        "Clash Description",
    "issue_titles":             "Issue Titles",
    "primary_concern":          "Primary Concern",
    "compliance_status":        "Compliance Status",
    "llm_certainty":            "AI Certainty",
    "constructability_assessment": "Constructability Assessment",
    "compliance_summary":       "Compliance Summary",
    "priority_band":            "Priority Band",
    "priority_score":           "Priority Score",
    "noise_filter_result":      "Noise Filter Result",
    "discipline_confidence":    "Discipline Classification Confidence",
    "recommended_action":       "Recommended Action",
    "suggested_solutions":      "Suggested Solutions",
    "requires_specialist":      "Requires Specialist?",
    "specialist_discipline":    "Specialist Discipline",
    "assigned_to":              "Assigned To",
    "assigned_email":           "Assigned Email",
    "response_required_by":     "Response Required By",
    "prepared_by":              "Prepared By",
    "status":                   "Status",
    "data_gaps":                "Data Gaps",
    "notes":                    "Notes",
}


# ── Helper ────────────────────────────────────────────────────────────────────

def _s(v, limit: Optional[int] = None) -> str:
    """Safe string — None → '', list → joined, dict → JSON."""
    if v is None:
        return ""
    if isinstance(v, list):
        v = " | ".join(str(x) for x in v if x not in (None, ""))
    elif isinstance(v, dict):
        v = json.dumps(v, default=str)
    s = str(v)
    return s[:limit] if limit else s


_LOCAL_SECRET_CACHE: Optional[dict] = None


def _read_local_secret_files() -> dict:
    """Fallback for local Streamlit secret files, including common accidental paths."""
    global _LOCAL_SECRET_CACHE
    if _LOCAL_SECRET_CACHE is not None:
        return _LOCAL_SECRET_CACHE

    try:
        import tomllib
    except Exception:
        _LOCAL_SECRET_CACHE = {}
        return _LOCAL_SECRET_CACHE

    root = Path(__file__).resolve().parent
    candidates = [
        root / ".streamlit" / "secrets.toml",
        root / "email" / ".streamlit" / "secrets.toml",
        root / "email..streamlit" / "secrets.toml",
    ]
    merged: dict = {}
    for candidate in candidates:
        if not candidate.exists():
            continue
        try:
            data = tomllib.loads(candidate.read_text())
        except Exception:
            continue
        for key, value in data.items():
            if key not in merged or str(key).upper().startswith("SMTP2GO"):
                merged[key] = value
    _LOCAL_SECRET_CACHE = merged
    return merged


def _lookup_nested(mapping: dict, dotted_key: str):
    cur: Any = mapping
    for part in dotted_key.split("."):
        if hasattr(cur, "get"):
            cur = cur.get(part)
        else:
            return None
        if cur is None:
            return None
    return cur


def _secret_value(*keys: str, default: str = "") -> str:
    """Read flat/nested Streamlit secrets, env vars, then local TOML fallbacks."""
    local_secrets = _read_local_secret_files()
    for key in keys:
        try:
            cur: Any = st.secrets
            for part in key.split("."):
                cur = cur.get(part) if hasattr(cur, "get") else cur[part]
                if cur is None:
                    break
            if cur not in (None, ""):
                return str(cur)
        except Exception:
            pass

        env_key = key.replace(".", "_").upper()
        env_val = os.environ.get(env_key)
        if env_val:
            return env_val

        local_val = _lookup_nested(local_secrets, key)
        if local_val in (None, "") and "." not in key:
            local_val = local_secrets.get(key)
        if local_val not in (None, ""):
            return str(local_val)

    return default


def _get_deepseek_key() -> str:
    return _secret_value("DEEPSEEK_API_KEY", "deepseek.api_key", default="")


DISCIPLINE_ALIASES = {
    "mechanical": "MECH", "mech": "MECH", "hvac": "MECH", "ventilation": "MECH",
    "electrical": "ELEC", "elec": "ELEC", "cable": "ELEC",
    "plumbing": "PLMB", "plumb": "PLMB", "pipe": "PLMB", "drainage": "PLMB", "water": "PLMB",
    "structural": "STR", "structure": "STR", "str": "STR", "beam": "STR", "column": "STR", "slab": "STR",
    "architecture": "ARCH", "architectural": "ARCH", "arch": "ARCH", "wall": "ARCH", "window": "ARCH", "door": "ARCH",
    "civil": "CIVIL", "site": "CIVIL",
    "fire protection": "FP", "fire safety": "FP", "fire": "FP", "sprinkler": "FP",
    "mep": "MECH",
}


def _discipline_label(code: str) -> str:
    return DISCIPLINES.get(code, {}).get("label", code or "Unassigned")


def _discipline_code(value) -> str:
    if value in (None, ""):
        return ""
    raw = str(value).strip()
    upper = raw.upper()

    for code in DISCIPLINES:
        if upper == code or upper.startswith(code + "-") or ("-" + code) in upper:
            return code

    lower = raw.lower()
    for alias, code in DISCIPLINE_ALIASES.items():
        if alias in lower:
            return code
    return ""


def _resolve_route_code(data: dict, cluster: Optional[dict] = None) -> str:
    cluster = cluster or {}
    issues = cluster.get("issues", []) or []
    first_issue = issues[0] if issues else {}
    disc_info = first_issue.get("_discipline") or {}

    candidates = [
        data.get("_route_discipline"),
        data.get("specialist_discipline"),
        data.get("_discipline"),
        data.get("discipline_label"),
        data.get("clash_pair"),
        cluster.get("discipline"),
        cluster.get("discipline_label"),
        cluster.get("clash_pair"),
        disc_info.get("primary"),
        disc_info.get("clash_pair"),
        data.get("cluster_label"),
        data.get("primary_concern"),
    ]
    for candidate in candidates:
        code = _discipline_code(candidate)
        if code:
            return DISC_EMAIL_MAP.get(code, code)
    return ""


def apply_prerfi_routing(data: dict, emails: dict, cluster: Optional[dict] = None) -> dict:
    routed = dict(data)
    route_code = _resolve_route_code(routed, cluster)
    routed["_route_discipline"] = route_code
    if route_code:
        routed["discipline_label"] = routed.get("discipline_label") or _discipline_label(route_code)
        routed["specialist_discipline"] = routed.get("specialist_discipline") or _discipline_label(route_code)
        routed["assigned_to"] = f"{_discipline_label(route_code)} Specialist"
        routed["assigned_email"] = emails.get(route_code, "")
    else:
        routed["assigned_to"] = routed.get("assigned_to") or "Unassigned Specialist"
        routed["assigned_email"] = ""
    return routed


def _group_prerfis_by_route(prerfi_list: list[dict], emails: dict) -> dict[str, list[dict]]:
    groups: dict[str, list[dict]] = {}
    for p in prerfi_list:
        d = p["data"]
        route_code = _resolve_route_code(d)
        recipient = emails.get(route_code, "") if route_code else ""
        d["_route_discipline"] = route_code
        d["assigned_email"] = recipient
        d["assigned_to"] = f"{_discipline_label(route_code)} Specialist" if route_code else "Unassigned Specialist"
        key = route_code if recipient else "__unassigned__"
        groups.setdefault(key, []).append(p)
    return groups


def _summarise_uploaded_template(uploaded) -> str:
    """Extract visible labels from an uploaded template so the LLM follows it."""
    if not uploaded:
        return ""

    name = getattr(uploaded, "name", "uploaded template")
    suffix = Path(name).suffix.lower()

    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook

            raw = uploaded.getvalue()
            wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            labels: list[str] = []
            max_row = min(ws.max_row or 1, 80)
            max_col = min(ws.max_column or 1, 12)
            for row in ws.iter_rows(max_row=max_row, max_col=max_col, values_only=True):
                for value in row:
                    label = _s(value).strip()
                    if label and label not in labels:
                        labels.append(label)
                    if len(labels) >= 90:
                        break
                if len(labels) >= 90:
                    break
            if labels:
                return f"Template file: {name}\nVisible template labels: " + " | ".join(labels)
        except Exception as exc:
            return f"Template file: {name} (could not parse XLSX labels: {exc})"

    return f"Template file: {name}; use the K1netix Pre-RFI sections and field labels."


def _fallback_solutions(data: dict) -> list[str]:
    refs = data.get("clause_references") or "the relevant project standard"
    return [
        f"Confirm the clash condition against {refs} and record whether a formal design response is required.",
        "Coordinate with the responsible discipline to propose a reroute, clearance adjustment, or element relocation.",
        "Update the BIM model and attach evidence screenshots before closing the Pre-RFI.",
    ]


def _apply_prerfi_fallback_narrative(data: dict) -> dict:
    """Fill weak blank fields with useful non-LLM drafting text."""
    d = dict(data)
    disc = d.get("discipline_label") or "the responsible discipline"
    label = d.get("cluster_label") or d.get("issue_titles") or "the clash group"
    refs = d.get("clause_references") or d.get("applicable_codes") or "retrieved regulation/project-standard context"
    status = d.get("compliance_status") or "NEEDS_REVIEW"
    priority = d.get("priority_band") or "MEDIUM"
    score = d.get("priority_score") or ""

    if not d.get("clash_description"):
        d["clash_description"] = (
            f"This Pre-RFI concerns {label}. The grouped issue involves {disc} coordination "
            f"and includes {d.get('cluster_size', 1)} source BCF issue(s). "
            f"Traceability is preserved through BCF topic GUID(s): {d.get('bcf_topic_guid') or 'not provided'}."
        )

    if not d.get("primary_concern"):
        d["primary_concern"] = (
            f"Potential {disc} coordination and compliance concern requiring human review "
            f"against {refs}."
        )

    if not d.get("constructability_assessment"):
        d["constructability_assessment"] = (
            "The issue may affect installation sequencing, access for inspection/maintenance, "
            "and the ability to construct the modelled elements without field rework. "
            "A discipline lead should confirm whether the modelled clearance and routing are acceptable."
        )

    if not d.get("compliance_summary"):
        quote = d.get("regulation_quote") or "No direct regulation quote was available from the retrieval step."
        d["compliance_summary"] = (
            f"Layer 3 classified this item as {status}. Retrieved reference context: {refs}. "
            f"Evidence summary: {quote}"
        )

    if not d.get("recommended_action"):
        d["recommended_action"] = (
            "Review the linked clash evidence, confirm the applicable requirement, and decide whether "
            "a design amendment, coordination meeting, or formal RFI escalation is needed."
        )

    if not d.get("suggested_solutions"):
        d["suggested_solutions"] = " | ".join(_fallback_solutions(d))

    if not d.get("specialist_discipline") and d.get("requires_specialist") == "Yes":
        d["specialist_discipline"] = disc

    if not d.get("data_gaps"):
        gaps = []
        if not d.get("floor_zone"):
            gaps.append("floor/zone not provided")
        if not d.get("element_a_guid") or not d.get("element_b_guid"):
            gaps.append("one or more BIM element GUIDs missing")
        if not d.get("regulation_quote"):
            gaps.append("direct regulation quote not retrieved")
        d["data_gaps"] = " | ".join(gaps) if gaps else "No major data gaps identified from available Layer 1-3 data."

    if not d.get("notes"):
        d["notes"] = (
            f"AI-generated advisory Pre-RFI. Priority {priority} with score {score}. "
            "Human review is required before this becomes a formal RFI."
        )

    return d


PRE_RFI_SYSTEM_PROMPT = """You are K1netix, a BIM coordination assistant drafting Pre-RFI forms for human review.

Write clear, professional construction coordination language. Use only the evidence provided.
Do not invent missing GUIDs, locations, regulation clauses, dates, or project IDs.
Preserve traceability fields exactly; only improve narrative fields.
Choose specialist_discipline from: Structural, Architecture, Mechanical, Electrical, Plumbing, Fire Protection, Civil.
Return valid JSON only, with concise but informative values suitable for an Excel Pre-RFI template."""


def _build_prerfi_prompt(data: dict, cluster: dict, template_summary: str = "") -> str:
    issues = cluster.get("issues", []) or []
    issue_context = []
    for idx, issue in enumerate(issues[:8], 1):
        issue_context.append({
            "index": idx,
            "title": issue.get("title", ""),
            "description": issue.get("description", ""),
            "status": issue.get("status", ""),
            "priority": issue.get("priority", ""),
            "guid": issue.get("guid") or issue.get("topic_guid") or "",
            "comments": issue.get("comments", [])[-2:] if issue.get("comments") else [],
        })

    payload = {
        "template_summary": template_summary[:2500],
        "pre_rfi_fields_already_extracted": {
            k: v for k, v in data.items() if not k.startswith("_")
        },
        "cluster_context": {
            "cluster_id": cluster.get("cluster_id", ""),
            "cluster_label": cluster.get("cluster_label", ""),
            "discipline": cluster.get("discipline", ""),
            "compliance": cluster.get("_compliance", {}),
            "priority_score": cluster.get("_priority_score", {}),
            "issues": issue_context,
        },
    }

    return (
        "Draft richer content for the narrative fields in this Pre-RFI form. "
        "Return JSON with these keys only: clash_description, primary_concern, "
        "constructability_assessment, compliance_summary, recommended_action, "
        "suggested_solutions, requires_specialist, specialist_discipline, "
        "life_safety_related, data_gaps, notes. suggested_solutions may be a list.\n\n"
        + json.dumps(payload, default=str, indent=2)
    )


def enhance_prerfi_with_llm(data: dict, cluster: dict, template_summary: str = "") -> dict:
    """Use DeepSeek to enrich the Pre-RFI narratives; fallback stays deterministic."""
    enriched = _apply_prerfi_fallback_narrative(data)
    api_key = _get_deepseek_key()

    if not api_key or DeepSeekClient is None:
        route_code = _resolve_route_code(enriched, cluster)
        enriched["_route_discipline"] = route_code
        if route_code and not enriched.get("specialist_discipline"):
            enriched["specialist_discipline"] = _discipline_label(route_code)
        enriched["_llm_drafting"] = "fallback"
        return enriched

    try:
        client = DeepSeekClient(api_key=api_key)
        draft = client.chat_json(
            _build_prerfi_prompt(enriched, cluster, template_summary),
            system_prompt=PRE_RFI_SYSTEM_PROMPT,
            temperature=0.2,
            max_tokens=1800,
        )

        allowed = {
            "clash_description", "primary_concern", "constructability_assessment",
            "compliance_summary", "recommended_action", "suggested_solutions",
            "requires_specialist", "specialist_discipline", "life_safety_related",
            "data_gaps", "notes",
        }
        for key in allowed:
            val = draft.get(key) if isinstance(draft, dict) else None
            if val in (None, "", []):
                continue
            if key == "suggested_solutions" and isinstance(val, list):
                enriched[key] = " | ".join(_s(x) for x in val if x not in (None, ""))
            elif key == "requires_specialist":
                enriched[key] = "Yes" if str(val).strip().lower() in {"yes", "true", "1", "required"} else "No"
            else:
                enriched[key] = _s(val)

        route_code = _resolve_route_code(enriched, cluster)
        enriched["_route_discipline"] = route_code
        if route_code and not enriched.get("specialist_discipline"):
            enriched["specialist_discipline"] = _discipline_label(route_code)
        enriched["_llm_drafting"] = "deepseek"
        return enriched
    except Exception as exc:
        route_code = _resolve_route_code(enriched, cluster)
        enriched["_route_discipline"] = route_code
        if route_code and not enriched.get("specialist_discipline"):
            enriched["specialist_discipline"] = _discipline_label(route_code)
        enriched["_llm_drafting"] = "fallback"
        existing = enriched.get("notes", "")
        warning = f"LLM drafting fallback used: {exc}"
        enriched["notes"] = f"{existing} | {warning}" if existing else warning
        return enriched


# ── Extract Pre-RFI data from cluster (pulls everything available) ────────────

def extract_prerfi_data(cluster: dict, rfi_number: str, project_name: str,
                        assigned_email: str = "",
                        project_location: str = "",
                        project_id: str = "",
                        bim_model_version: str = "",
                        clash_source_tool: str = "Navisworks") -> dict:
    comp   = cluster.get("_compliance") or {}
    ps     = cluster.get("_priority_score") or {}
    issues = cluster.get("issues", []) or []

    first_issue = issues[0] if issues else {}
    disc_info   = first_issue.get("_discipline") or {}
    noise       = first_issue.get("_noise_filter") or {}

    # BCF / element extraction — try multiple possible field names from Layer 1
    bcf_guids = [i.get("guid") or i.get("topic_guid") or "" for i in issues]
    bcf_guid_str = " | ".join(g for g in bcf_guids if g)[:200]

    # Element data — try common BCF field names
    el_a_guid = (first_issue.get("element_a_guid")
                 or first_issue.get("component_a")
                 or first_issue.get("element_a")
                 or "")
    el_b_guid = (first_issue.get("element_b_guid")
                 or first_issue.get("component_b")
                 or first_issue.get("element_b")
                 or "")
    el_a_type = first_issue.get("element_a_type","")
    el_b_type = first_issue.get("element_b_type","")

    # Coordinates — try multiple shapes
    coords = (first_issue.get("clash_coordinates")
              or first_issue.get("coordinates")
              or first_issue.get("location")
              or "")
    if isinstance(coords, dict):
        coords = f"({coords.get('x','?')}, {coords.get('y','?')}, {coords.get('z','?')})"
    elif isinstance(coords, (list, tuple)) and len(coords) == 3:
        coords = f"({coords[0]}, {coords[1]}, {coords[2]})"

    # Floor / zone
    floor_zone = (first_issue.get("floor")
                  or first_issue.get("zone")
                  or first_issue.get("level")
                  or "")

    # Layer 3 — regulation data
    clause_checks = comp.get("clause_checks") or []
    clause_refs   = " | ".join(
        f"{c.get('clause_ref','')} [{c.get('status','UNCERTAIN')}]"
        for c in clause_checks if c.get("clause_ref")
    )
    codes = " | ".join(sorted({
        c.get("clause_ref","").split("§")[0].strip()
        for c in clause_checks if c.get("clause_ref")
    }))
    reg_quote = " | ".join(
        _s(c.get("clause_summary",""), 200)
        for c in clause_checks[:3] if c.get("clause_summary")
    )

    # Discipline
    disc_code  = cluster.get("discipline","") or disc_info.get("primary", "") or ""
    disc_label = cluster.get("discipline_label") or DISCIPLINES.get(disc_code, {}).get("label", disc_code)
    disc_conf  = disc_info.get("confidence", 0) or 0
    disc_conf_str = f"{round(disc_conf * 100, 1)}%" if disc_conf else ""

    # Solutions and gaps
    solutions = comp.get("suggested_solutions") or []
    data_gaps = comp.get("data_gaps") or []

    # Issues summary
    issue_titles = " | ".join(_s(i.get("title","")) for i in issues[:8])
    clash_desc   = " | ".join(_s(i.get("description",""), 150)
                              for i in issues[:3] if i.get("description"))

    # Response required by — 14 days from now for default
    response_due = (datetime.now() + timedelta(days=14)).strftime("%d/%m/%Y")

    composite = ps.get("composite") or 0
    life_safety = "Yes" if composite > 0.7 else ("Possibly" if composite > 0.4 else "No")

    return {
        # Section 1
        "project_name":             _s(project_name),
        "project_location":         _s(project_location),
        "discipline_label":         disc_label,
        "project_id":               _s(project_id),
        "bim_model_version":        _s(bim_model_version),
        "date_generated":           datetime.now().strftime("%d/%m/%Y"),
        "rfi_number":               rfi_number,

        # Section 2
        "traceability_id":          _s(cluster.get("cluster_id","")),
        "bcf_topic_guid":           bcf_guid_str or _s(first_issue.get("guid","")),
        "clash_source_tool":        clash_source_tool,
        "floor_zone":               _s(floor_zone),
        "element_a_guid":           _s(el_a_guid),
        "element_a_type":           _s(el_a_type),
        "element_b_guid":           _s(el_b_guid),
        "element_b_type":           _s(el_b_type),
        "clash_coordinates":        _s(coords),
        "cluster_size":             len(issues),
        "cluster_label":            _s(cluster.get("cluster_label","")),
        "clash_pair":               _s(disc_info.get("clash_pair","") or cluster.get("clash_pair", "")),

        # Section 3
        "applicable_codes":         codes,
        "clause_references":        clause_refs,
        "regulation_version":       "",  # filled if available in clause meta
        "regulation_quote":         reg_quote,
        "life_safety_related":      life_safety,
        "regulation_match_quality": f"{round((comp.get('regulation_match_quality') or 0)*100)}%",

        # Section 4
        "clash_description":        clash_desc,
        "issue_titles":             issue_titles,
        "primary_concern":          _s(comp.get("primary_concern","")),
        "compliance_status":        _s(comp.get("compliance_status","INSUFFICIENT_DATA")),
        "llm_certainty":            f"{round((comp.get('llm_certainty') or 0)*100)}%",
        "constructability_assessment": _s(comp.get("constructability_assessment","")),
        "compliance_summary":       _s(comp.get("summary","")),
        "priority_band":            _s(ps.get("band","MEDIUM")),
        "priority_score":           round(composite, 4),
        "noise_filter_result":      _s(noise.get("decision","")),
        "discipline_confidence":    disc_conf_str,

        # Section 5
        "recommended_action":       _s(comp.get("recommended_action","")),
        "suggested_solutions":      _s(solutions),
        "requires_specialist":      "Yes" if comp.get("requires_specialist") else "No",
        "specialist_discipline":    _s(comp.get("specialist_discipline","")),

        # Section 6
        "assigned_to":              disc_label + " Engineer",
        "assigned_email":           assigned_email,
        "response_required_by":     response_due,
        "prepared_by":              "K1netix AI",
        "status":                   "Open · Awaiting Review",
        "data_gaps":                _s(data_gaps),
        "notes":                    "",

        # Internal
        "_discipline":              disc_code,
        "_cluster_id":              _s(cluster.get("cluster_id","")),
        "_generated_at":            datetime.now().isoformat(),
    }


# ── Excel builder ─────────────────────────────────────────────────────────────

def build_xlsx_output(prerfi_list: list[dict], template_bytes: Optional[bytes] = None) -> bytes:
    """Build an Excel package whose detailed sheets follow the Pre-RFI form template."""
    from copy import copy
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    STATUS_COLORS = {
        "NON_COMPLIANT":    "ff6b6b",
        "NEEDS_REVIEW":     "ffd060",
        "COMPLIANT":        "6bff9e",
        "INSUFFICIENT_DATA":"cccccc",
    }
    PRIORITY_COLORS = {
        "CRITICAL": "ff4444", "HIGH": "ff9900",
        "MEDIUM":   "ffd060", "LOW":  "88cc88",
    }

    def _safe_sheet_name(name: str, fallback: str) -> str:
        cleaned = "".join(ch for ch in (name or fallback) if ch not in r'[]:*?/\\')[:31]
        return cleaned or fallback[:31]

    def _to_score(value, default: float = 0.0) -> float:
        if value in (None, ""):
            return default
        if isinstance(value, bool):
            return 1.0 if value else 0.0
        if isinstance(value, (int, float)):
            return max(0.0, min(1.0, float(value)))
        s = str(value).strip().replace("%", "")
        try:
            n = float(s)
            if n > 1:
                n = n / 100
            return max(0.0, min(1.0, n))
        except Exception:
            return default

    def _data_completeness(d: dict) -> float:
        important = [
            "bcf_topic_guid", "cluster_label", "discipline_label", "clash_pair",
            "clash_description", "primary_concern", "compliance_summary",
            "recommended_action", "suggested_solutions", "clause_references",
        ]
        present = sum(1 for key in important if _s(d.get(key)).strip())
        return round(present / len(important), 2)

    def _load_template_wb() -> Workbook:
        if template_bytes:
            return load_workbook(io.BytesIO(template_bytes))

        default_template = Path('/Users/qitia/Desktop/AI Challenge/Copy of K1netix_Pre_RFI_Template.xlsx')
        if default_template.exists():
            return load_workbook(default_template)

        wb = Workbook()
        ws = wb.active
        ws.title = "Pre-RFI Form"
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = [3, 26, 22, 18, 22, 18, 18][col-1]
        blue = PatternFill("solid", fgColor="305c9b")
        pale = PatternFill("solid", fgColor="fff2cc")
        label = PatternFill("solid", fgColor="ddebf7")
        thin = Side(style="thin", color="9eadd6")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        def merge(range_ref, value="", fill=None, color="000000", bold=False, size=11, italic=False, align="left"):
            ws.merge_cells(range_ref)
            cell = ws[range_ref.split(":")[0]]
            cell.value = value
            cell.fill = fill or PatternFill(fill_type=None)
            cell.font = Font(name="Arial", color=color, bold=bold, size=size, italic=italic)
            cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
            cell.border = border
        merge("B2:G2", "PRE-RFI DOCUMENT", color="1f3b63", bold=True, size=18, align="center")
        merge("B3:G3", "K1netix -- AI-Generated Advisory Document for Human Review", color="4472c4", italic=True, align="center")
        for row, title in [(6,"1. PROJECT INFORMATION"),(12,"2. TRACEABILITY & CLASH IDENTIFICATION"),(20,"3. REGULATION & COMPLIANCE REFERENCE"),(29,"4. AI REASONING CHAIN"),(39,"5. PROPOSED SOLUTION"),(47,"6. PRIORITIZATION SCORE"),(56,"7. HUMAN REVIEW & SIGN-OFF")]:
            merge(f"B{row}:G{row}", title, fill=blue, color="ffffff", bold=True)
        label_cells = {
            "B7":"Project Name", "B8":"Project Location", "E8":"Project ID", "B9":"Discipline(s)", "E9":"BIM Model Version", "B10":"Date Generated", "E10":"Pre-RFI Number",
            "B13":"Traceability ID", "E13":"BCF Topic GUID", "B14":"Clash Source Tool", "E14":"Floor / Zone", "B15":"Element A (GUID)", "E15":"Element A Type", "B16":"Element B (GUID)", "E16":"Element B Type", "B17":"Clash Coordinates (X, Y, Z)", "E17":"Cluster Size",
            "B21":"Applicable Code / Standard", "B22":"Clause / Section Number", "F22":"Version / Year", "B23":"Regulation Quote", "B27":"Life-Safety Related?", "E27":"Regulation Match Quality",
            "B30":"Clash Description", "B34":"Non-Compliance Reasoning", "B40":"Recommended Action", "B44":"Estimated Cost Impact", "E44":"Amount ($)", "B45":"Estimated Time Impact", "E45":"# of Days",
            "B48":"Component", "D48":"Weight", "E48":"Score (0-1)", "F48":"Weighted", "B49":"LLM Confidence Score", "B50":"Input Data Completeness", "B51":"Classifier Prediction Probability", "B52":"Regulation Match Quality", "B53":"COMPOSITE SCORE", "B54":"Priority Level", "B57":"Reviewer Name & Title", "E57":"Review Date", "B58":"Reviewer Email", "E58":"Phone", "B59":"Review Decision", "B60":"Reviewer Comments",
        }
        for cell, value in label_cells.items():
            ws[cell] = value
            ws[cell].font = Font(name="Arial", bold=True, color="1f3b63", size=10)
            ws[cell].fill = label
            ws[cell].border = border
            ws[cell].alignment = Alignment(wrap_text=True, vertical="center")
        for rng in ["B7:C7","D7:G7","C8:D8","F8:G8","C9:D9","F9:G9","C10:D10","F10:G10","C13:D13","F13:G13","C14:D14","F14:G14","C15:D15","F15:G15","C16:D16","F16:G16","C17:D17","F17:G17","B18:G18","C21:G21","C22:E22","F22:G22","B24:G26","C27:D27","F27:G27","B31:G33","B35:G37","B41:G43","C44:D44","F44:G44","C45:D45","F45:G45","B48:C48","F48:G48","B49:C49","F49:G49","B50:C50","F50:G50","B51:C51","F51:G51","B52:C52","F52:G52","B53:E53","F53:G53","C54:G54","C57:D57","F57:G57","C58:D58","F58:G58","C59:G59","B61:G63","B65:G65"]:
            if rng not in [str(r) for r in ws.merged_cells.ranges]:
                ws.merge_cells(rng)
        for row in range(1, 66):
            ws.row_dimensions[row].height = 22
        for row in [24,25,26,31,32,33,35,36,37,41,42,43,61,62,63]:
            ws.row_dimensions[row].height = 34
        for row in range(1, 66):
            for col in range(2, 8):
                c = ws.cell(row, col)
                c.border = border
                c.alignment = Alignment(wrap_text=True, vertical="top")
        merge("B65:G65", "Generated by K1netix -- AI-Driven Clash Triage, Compliance Reasoning, and Pre-RFI Generation", color="666666", italic=True, align="center")
        return wb

    def _write(cell, value):
        cell.value = value
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    def _fill_form(ws, d: dict):
        clash_description = "\n".join(x for x in [
            d.get("clash_description", ""),
            f"Source issue title(s): {d.get('issue_titles','')}",
            f"Affected elements: {d.get('element_a_type','Element A')} {d.get('element_a_guid','')} / {d.get('element_b_type','Element B')} {d.get('element_b_guid','')}",
            f"Location: {d.get('floor_zone','not provided')} | Coordinates: {d.get('clash_coordinates','not provided')}",
            f"Source tool: {d.get('clash_source_tool','')} | BCF GUID(s): {d.get('bcf_topic_guid','')}",
        ] if _s(x).strip() and not _s(x).endswith(': '))

        reasoning = "\n\n".join(x for x in [
            f"Primary concern: {d.get('primary_concern','')}",
            f"Compliance status: {d.get('compliance_status','')} | AI certainty: {d.get('llm_certainty','')} | Regulation match quality: {d.get('regulation_match_quality','')}",
            f"Applicable code/standard: {d.get('applicable_codes','')}",
            f"Clause/section reference: {d.get('clause_references','')}",
            f"Regulation quote/evidence: {d.get('regulation_quote','')}",
            f"Compliance summary: {d.get('compliance_summary','')}",
            f"Constructability assessment: {d.get('constructability_assessment','')}",
            f"Life-safety related: {d.get('life_safety_related','')}",
        ] if _s(x).strip() and not _s(x).endswith(': '))

        solution = "\n\n".join(x for x in [
            f"Recommended action: {d.get('recommended_action','')}",
            f"Suggested solution option(s): {d.get('suggested_solutions','')}",
            f"Responsible specialist: {d.get('specialist_discipline') or d.get('discipline_label','')}",
            f"Assigned contact: {d.get('assigned_email','not assigned')}",
            f"Response required by: {d.get('response_required_by','')}",
            f"Data gaps to verify: {d.get('data_gaps','')}",
        ] if _s(x).strip() and not _s(x).endswith(': '))

        reviewer_packet = "\n".join(x for x in [
            f"Prepared by: {d.get('prepared_by','K1netix AI')}",
            f"Routing discipline: {_discipline_label(d.get('_route_discipline',''))}",
            f"Assigned to: {d.get('assigned_to','')}",
            f"Assigned email: {d.get('assigned_email','')}",
            f"Priority: {d.get('priority_band','')} (score {d.get('priority_score','')})",
            f"Noise filter result: {d.get('noise_filter_result','')}",
            f"Discipline confidence: {d.get('discipline_confidence','')}",
            f"Traceability ID: {d.get('traceability_id','')}",
            f"Cluster label: {d.get('cluster_label','')}",
            f"Clash pair: {d.get('clash_pair','')}",
            f"Notes: {d.get('notes','')}",
        ] if _s(x).strip() and not _s(x).endswith(': '))

        values = {
            "D7": d.get("project_name", ""),
            "C8": d.get("project_location", ""),
            "F8": d.get("project_id", ""),
            "C9": d.get("discipline_label", ""),
            "F9": d.get("bim_model_version", ""),
            "C10": d.get("date_generated", ""),
            "F10": d.get("rfi_number", ""),
            "C13": d.get("traceability_id", ""),
            "F13": d.get("bcf_topic_guid", ""),
            "C14": d.get("clash_source_tool", ""),
            "F14": d.get("floor_zone", ""),
            "C15": d.get("element_a_guid", ""),
            "F15": d.get("element_a_type", ""),
            "C16": d.get("element_b_guid", ""),
            "F16": d.get("element_b_type", ""),
            "C17": d.get("clash_coordinates", ""),
            "F17": d.get("cluster_size", ""),
            "C21": d.get("applicable_codes", ""),
            "C22": d.get("clause_references", ""),
            "F22": d.get("regulation_version", ""),
            "B24": d.get("regulation_quote", ""),
            "C27": "Yes" if str(d.get("life_safety_related", "")).lower() in {"yes", "true", "possibly"} else "No",
            "F27": d.get("regulation_match_quality", ""),
            "B31": clash_description,
            "B35": reasoning,
            "B41": solution,
            "C44": "No Change / TBC",
            "F44": "TBC",
            "C45": "No Change / TBC",
            "F45": "TBC",
            "D49": 0.25,
            "E49": _to_score(d.get("llm_certainty"), 0.5),
            "D50": 0.25,
            "E50": _data_completeness(d),
            "D51": 0.25,
            "E51": _to_score(d.get("discipline_confidence"), 0.5),
            "D52": 0.25,
            "E52": _to_score(d.get("regulation_match_quality"), 0.1),
            "C54": f"{d.get('priority_band','MEDIUM')} (composite score {d.get('priority_score','')})",
            "C57": "",
            "F57": "",
            "C58": d.get("assigned_email", ""),
            "F58": "",
            "C59": "Requires Human Review",
            "B61": reviewer_packet,
            "B65": f"Generated by K1netix -- {datetime.now().strftime('%d %b %Y %H:%M')} -- AI advisory document for human review",
        }
        for ref, value in values.items():
            _write(ws[ref], _s(value))

        ws["B18"] = f"Note: Cluster Size = {d.get('cluster_size', 1)} related clash(es). Group: {d.get('cluster_label','')}. Clash pair: {d.get('clash_pair','')}."
        ws["F49"] = "=D49*E49"
        ws["F50"] = "=D50*E50"
        ws["F51"] = "=D51*E51"
        ws["F52"] = "=D52*E52"
        ws["F53"] = "=SUM(F49:F52)"

        status = d.get("compliance_status", "")
        priority = d.get("priority_band", "")
        status_color = STATUS_COLORS.get(status, "fff2cc")
        priority_color = PRIORITY_COLORS.get(priority, "fff2cc")
        ws["C59"].fill = PatternFill("solid", fgColor=status_color)
        ws["C54"].fill = PatternFill("solid", fgColor=priority_color)

        for ref in ["D7","C8","F8","C9","F9","C10","F10","C13","F13","C14","F14","C15","F15","C16","F16","C17","F17","C21","C22","F22","B24","C27","F27","B31","B35","B41","C44","F44","C45","F45","D49","E49","D50","E50","D51","E51","D52","E52","C54","C57","F57","C58","F58","C59","B61"]:
            if ws[ref].fill.fill_type != "solid":
                ws[ref].fill = PatternFill("solid", fgColor="fff2cc")

        row_heights = {
            24: 42, 25: 42, 26: 42,
            31: 68, 32: 68, 33: 68,
            35: 82, 36: 82, 37: 82,
            41: 74, 42: 74, 43: 74,
            61: 72, 62: 72, 63: 72,
        }
        for row, height in row_heights.items():
            ws.row_dimensions[row].height = height

    wb = _load_template_wb()
    base = wb.active
    for ws in list(wb.worksheets):
        if ws is not base:
            wb.remove(ws)

    form_sheets = [base]
    for _ in prerfi_list[1:]:
        form_sheets.append(wb.copy_worksheet(base))

    used_names = set()
    for idx, (ws, p) in enumerate(zip(form_sheets, prerfi_list), start=1):
        d = p["data"]
        name = _safe_sheet_name(d.get("rfi_number", f"Pre-RFI {idx}"), f"Pre-RFI {idx}")
        if name in used_names:
            name = _safe_sheet_name(f"{name}-{idx}", f"Pre-RFI {idx}")
        used_names.add(name)
        ws.title = name
        _fill_form(ws, d)

    # Keep the filled form sheets first, so the workbook opens like the uploaded template.
    ws_reg = wb.create_sheet("Pre-RFI Register")
    ws_reg.freeze_panes = "A3"
    headers = ["#", "RFI Number", "Status", "Discipline", "Assigned Email", "Priority", "Cluster", "Recommended Action", "Response Due"]
    ws_reg.merge_cells("A1:I1")
    ws_reg["A1"] = f"K1netix Pre-RFI Register - Generated {datetime.now().strftime('%d %b %Y %H:%M')}"
    ws_reg["A1"].font = Font(name="Arial", bold=True, size=13, color="f0c040")
    ws_reg["A1"].fill = PatternFill("solid", fgColor="0f1117")
    ws_reg["A1"].alignment = Alignment(horizontal="center")
    for col, header in enumerate(headers, 1):
        c = ws_reg.cell(2, col, header)
        c.font = Font(name="Arial", bold=True, color="ffffff")
        c.fill = PatternFill("solid", fgColor="1a1d27")
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    widths = [6, 18, 18, 16, 30, 12, 36, 52, 16]
    for col, width in enumerate(widths, 1):
        ws_reg.column_dimensions[get_column_letter(col)].width = width
    thin = Side(style="thin", color="bbbbbb")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for idx, p in enumerate(prerfi_list, 1):
        d = p["data"]
        values = [
            idx, d.get("rfi_number", ""), d.get("compliance_status", ""),
            d.get("discipline_label", ""), d.get("assigned_email", ""),
            d.get("priority_band", ""), d.get("cluster_label", ""),
            d.get("recommended_action", ""), d.get("response_required_by", ""),
        ]
        for col, value in enumerate(values, 1):
            c = ws_reg.cell(idx + 2, col, value)
            c.border = border
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if col == 3:
                c.fill = PatternFill("solid", fgColor=STATUS_COLORS.get(d.get("compliance_status", ""), "ffffff"))
            elif col == 6:
                c.fill = PatternFill("solid", fgColor=PRIORITY_COLORS.get(d.get("priority_band", ""), "ffffff"))
        ws_reg.row_dimensions[idx + 2].height = 42

    ws_app = wb.create_sheet("Evidence Appendix")
    app_headers = ["RFI Number", "Field", "Value"]
    for col, header in enumerate(app_headers, 1):
        c = ws_app.cell(1, col, header)
        c.font = Font(name="Arial", bold=True, color="ffffff")
        c.fill = PatternFill("solid", fgColor="1a1d27")
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    ws_app.column_dimensions["A"].width = 18
    ws_app.column_dimensions["B"].width = 30
    ws_app.column_dimensions["C"].width = 100
    app_row = 2
    appendix_skip = {"_generated_at"}
    for p in prerfi_list:
        d = p["data"]
        for key, value in d.items():
            if key in appendix_skip:
                continue
            ws_app.cell(app_row, 1, d.get("rfi_number", ""))
            ws_app.cell(app_row, 2, FIELD_LABELS.get(key, key))
            ws_app.cell(app_row, 3, _s(value))
            for col in range(1, 4):
                c = ws_app.cell(app_row, col)
                c.border = border
                c.alignment = Alignment(wrap_text=True, vertical="top")
            ws_app.row_dimensions[app_row].height = 34 if len(_s(value)) > 100 else 20
            app_row += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── SMTP from secrets ─────────────────────────────────────────────────────────

def _get_smtp_config() -> dict:
    """Read SMTP2GO/SMTP credentials from Streamlit secrets or environment."""
    smtp2go_user = _secret_value(
        "SMTP2GO_USER", "SMTP2GO_USERNAME", "SMTP2GO_SMTP_USERNAME",
        "smtp2go.user", "smtp2go.username", "smtp2go.smtp_username",
        default="",
    )
    smtp2go_password = _secret_value(
        "SMTP2GO_PASSWORD", "SMTP2GO_PASS", "SMTP2GO_SMTP_PASSWORD",
        "smtp2go.password", "smtp2go.pass", "smtp2go.smtp_password",
        default="",
    )

    if smtp2go_user or smtp2go_password:
        host = _secret_value("SMTP2GO_HOST", "smtp2go.host", default="mail.smtp2go.com")
        port_raw = _secret_value("SMTP2GO_PORT", "smtp2go.port", default="2525")
        user = smtp2go_user
        password = smtp2go_password
        provider = "SMTP2GO"
    else:
        host = _secret_value("SMTP_HOST", "smtp.host", default="mail.smtp2go.com")
        port_raw = _secret_value("SMTP_PORT", "smtp.port", default="2525")
        user = _secret_value("SMTP_USER", "SMTP_USERNAME", "smtp.user", "smtp.username", default="")
        password = _secret_value("SMTP_PASSWORD", "SMTP_PASS", "smtp.password", "smtp.pass", default="")
        provider = "SMTP2GO" if "smtp2go" in host.lower() else "SMTP"

    try:
        port = int(port_raw)
    except Exception:
        port = 2525

    from_email = _secret_value(
        "SMTP2GO_FROM", "SMTP2GO_FROM_EMAIL", "SMTP2GO_SENDER", "smtp2go.from_email", "smtp2go.sender",
        "SMTP_FROM", "SMTP_FROM_EMAIL", "SMTP_SENDER", "SENDER_EMAIL", "smtp.from_email", "smtp.sender",
        default=user,
    )
    from_name = _secret_value(
        "SMTP2GO_FROM_NAME", "smtp2go.from_name",
        "SMTP_FROM_NAME", "smtp.from_name",
        default="K1netix AI",
    )
    use_ssl = _secret_value(
        "SMTP_USE_SSL", "smtp.use_ssl", "SMTP2GO_USE_SSL", "smtp2go.use_ssl",
        default="true" if port == 465 else "false",
    ).lower() in {"1", "true", "yes", "on"}
    starttls = _secret_value(
        "SMTP_STARTTLS", "smtp.starttls", "SMTP2GO_STARTTLS", "smtp2go.starttls",
        default="false" if port == 465 else "true",
    ).lower() in {"1", "true", "yes", "on"}

    return {
        "host": host,
        "port": port,
        "user": user,
        "password": password,
        "from_email": from_email,
        "from_name": from_name,
        "use_ssl": use_ssl,
        "starttls": starttls,
        "provider": provider,
    }


def _make_ssl_context() -> ssl.SSLContext:
    """SSL context with certifi cert bundle to fix macOS cert issues."""
    if _CERT_PATH:
        return ssl.create_default_context(cafile=_CERT_PATH)
    return ssl.create_default_context()


def send_email(recipient_email: str, subject: str, body_html: str,
               attachment_bytes: Optional[bytes] = None,
               attachment_name: str = "PreRFI.xlsx") -> dict:
    cfg = _get_smtp_config()
    sender = cfg.get("from_email") or cfg.get("user")
    if not cfg["user"] or not cfg["password"] or not sender:
        return {"ok": False, "error": "SMTP2GO credentials or sender email missing from Streamlit secrets"}

    try:
        msg = MIMEMultipart("mixed")
        msg["From"]    = formataddr((cfg.get("from_name", "K1netix"), sender))
        msg["To"]      = recipient_email
        msg["Subject"] = subject
        msg.attach(MIMEText(body_html, "html"))

        if attachment_bytes:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(attachment_bytes)
            encoders.encode_base64(part)
            part.add_header("Content-Disposition",
                            f'attachment; filename="{attachment_name}"')
            msg.attach(part)

        def _do_send(ctx):
            if cfg["use_ssl"] or cfg["port"] == 465:
                with smtplib.SMTP_SSL(cfg["host"], cfg["port"], context=ctx) as s:
                    s.login(cfg["user"], cfg["password"])
                    s.sendmail(sender, recipient_email, msg.as_string())
            else:
                with smtplib.SMTP(cfg["host"], cfg["port"]) as s:
                    s.ehlo()
                    if cfg.get("starttls", True):
                        s.starttls(context=ctx)
                        s.ehlo()
                    s.login(cfg["user"], cfg["password"])
                    s.sendmail(sender, recipient_email, msg.as_string())

        # Try contexts in order: certifi → default → no-verify (macOS workaround)
        try:
            _do_send(_make_ssl_context())
        except (ssl.SSLError, ssl.SSLCertVerificationError) as ssl_err:
            # Last resort: disable cert verification (still encrypted, just unverified)
            unverified = ssl.create_default_context()
            unverified.check_hostname = False
            unverified.verify_mode    = ssl.CERT_NONE
            _do_send(unverified)

        return {"ok": True}
    except Exception as e:
        return {"ok": False, "error": str(e)}


def _email_body(group: list[dict], disc_label: str) -> str:
    rows = ""
    STATUS_COLORS_HEX = {
        "NON_COMPLIANT":"#ff6b6b","NEEDS_REVIEW":"#ffd060",
        "COMPLIANT":"#6bff9e","INSUFFICIENT_DATA":"#cccccc",
    }
    rfi_numbers = ", ".join(p["data"].get("rfi_number", "") for p in group)
    for p in group:
        d      = p["data"]
        status = d.get("compliance_status","")
        color  = STATUS_COLORS_HEX.get(status,"#eee")
        rows  += f"""<tr>
          <td style="padding:8px;border:1px solid #ddd">{d.get('rfi_number','')}</td>
          <td style="padding:8px;border:1px solid #ddd">{d.get('cluster_label','')[:70]}</td>
          <td style="padding:8px;border:1px solid #ddd;background:{color};font-weight:bold">{status}</td>
          <td style="padding:8px;border:1px solid #ddd">{d.get('priority_band','')}</td>
          <td style="padding:8px;border:1px solid #ddd">{d.get('primary_concern','')[:160]}</td>
          <td style="padding:8px;border:1px solid #ddd">{d.get('recommended_action','')[:160]}</td>
        </tr>"""

    return f"""<html><body>
<div style="font-family:Arial,sans-serif;max-width:950px;margin:0 auto;color:#1a1d27">
  <div style="background:#0f1117;padding:24px;border-radius:8px 8px 0 0">
    <h1 style="color:#f0c040;font-family:monospace;margin:0">K1netix AI</h1>
    <p style="color:#c9ced8;margin:4px 0 0">BIM Coordination Intelligence - Pre-RFI Package</p>
  </div>
  <div style="background:#fff;padding:24px;border:1px solid #eee;border-top:none">
    <p>Hello,</p>
    <p>This is K1netix AI. A Pre-RFI package has been generated for your review because the BIM coordination workflow identified clash issue(s) assigned to <strong>{disc_label}</strong>.</p>
    <p>Please open the attached Excel file and review the full Pre-RFI form sheet(s). The document includes traceability IDs, BCF topic GUIDs, affected elements, retrieved regulation references, AI reasoning, recommended actions, and human-review fields.</p>
    <p><strong>Included Pre-RFI(s):</strong> {rfi_numbers}</p>
    <table style="width:100%;border-collapse:collapse;margin-top:16px;font-size:13px">
      <thead><tr style="background:#1a3a6b;color:#fff">
        <th style="padding:10px;text-align:left">RFI #</th>
        <th style="padding:10px;text-align:left">Clash Group</th>
        <th style="padding:10px;text-align:left">Status</th>
        <th style="padding:10px;text-align:left">Priority</th>
        <th style="padding:10px;text-align:left">Primary Concern</th>
        <th style="padding:10px;text-align:left">Recommended Action</th>
      </tr></thead>
      <tbody>{rows}</tbody>
    </table>
    <p style="margin-top:24px;color:#666;font-size:12px">
      This is an AI-generated advisory document. Please review and approve before issuing any formal RFI or design instruction.<br>
      Generated by K1netix AI - {datetime.now().strftime("%d %b %Y %H:%M")}
    </p>
  </div>
</div></body></html>"""


# ── Record store ──────────────────────────────────────────────────────────────

def _init_records():
    if "layer4_records" not in st.session_state:
        st.session_state["layer4_records"] = []


# ── Main Streamlit UI ─────────────────────────────────────────────────────────

def render_layer4_ui(layer3_result: dict):
    _init_records()

    st.markdown("---")
    st.markdown("## 📋 Layer 4 — Pre-RFI Generator & Router")
    st.caption("Enter contacts first → generate template-style Pre-RFIs → route and send in one workflow")

    clusters = layer3_result.get("clusters", [])
    if not clusters:
        st.info("Run Layer 3 compliance check first to generate Pre-RFIs.")
        return

    actionable = [c for c in clusters
                  if c.get("_compliance", {}).get("compliance_status")
                  in ("NON_COMPLIANT", "NEEDS_REVIEW")]

    n_skip = len(clusters) - len(actionable)
    st.markdown(f"**{len(actionable)}** clusters require Pre-RFIs "
                f"({n_skip} compliant/skipped)")

    tabs = st.tabs(["🧭 Workflow", "📒 Records"])

    with tabs[0]:
        default_proj_id = f"PRJ_{datetime.now().strftime('%Y_%m')}_001"

        with st.expander("1. Project Settings", expanded=False):
            st.caption("These fields populate the header cells in the Pre-RFI form.")
            c1, c2 = st.columns(2)
            with c1:
                project_name = st.text_input(
                    "Project Name",
                    value=st.session_state.get("layer4_project_name", "K1netix Project"),
                    key="l4_proj_name",
                )
                project_id = st.text_input(
                    "Project ID",
                    value=st.session_state.get("layer4_project_id", default_proj_id),
                    key="l4_proj_id",
                )
                clash_source = st.text_input(
                    "Clash Source Tool",
                    value=st.session_state.get("layer4_clash_tool", "Navisworks"),
                    key="l4_clash_tool",
                )
            with c2:
                project_location = st.text_input(
                    "Project Location",
                    value=st.session_state.get("layer4_proj_loc", "—"),
                    key="l4_proj_loc",
                )
                bim_version = st.text_input(
                    "BIM Model Version",
                    value=st.session_state.get("layer4_bim_ver", f"v1 — {datetime.now().strftime('%d %b %Y')}"),
                    key="l4_bim_ver",
                )
                prepared_by = st.text_input(
                    "Prepared By",
                    value=st.session_state.get("layer4_prep_by", "K1netix AI"),
                    key="l4_prep_by",
                )

        st.session_state.update({
            "layer4_project_name": project_name,
            "layer4_project_id": project_id,
            "layer4_clash_tool": clash_source,
            "layer4_proj_loc": project_location,
            "layer4_bim_ver": bim_version,
            "layer4_prep_by": prepared_by,
        })

        with st.expander("2. Pre-RFI Template", expanded=False):
            st.caption(
                "Upload your Excel Pre-RFI template. The generated workbook keeps the form sheet first "
                "and moves registers/appendix sheets behind it."
            )
            uploaded = st.file_uploader(
                "Upload Pre-RFI template (.xlsx)",
                type=["xlsx"],
                key="l4_template_upload",
            )
            if uploaded:
                st.session_state["layer4_custom_template_name"] = uploaded.name
                st.session_state["layer4_template_bytes"] = uploaded.getvalue()
                st.session_state["layer4_template_summary"] = _summarise_uploaded_template(uploaded)
                st.success(f"Template loaded: **{uploaded.name}**")
            elif st.session_state.get("layer4_template_bytes"):
                st.success(f"Template still loaded: **{st.session_state.get('layer4_custom_template_name','uploaded template')}**")
            else:
                st.info("No template upload detected. K1netix will use the local/default Pre-RFI form layout if available.")

        st.markdown("### 3. Specialist Contacts")
        st.caption("Fill contacts before generation so each Pre-RFI is created with its routing email already attached.")
        emails = st.session_state.get("layer4_emails", {d: "" for d in DISCIPLINES})
        contact_cols = st.columns(2)
        idx = 0
        for disc_code, disc_cfg in DISCIPLINES.items():
            if disc_code == "MEP":
                continue
            with contact_cols[idx % 2]:
                emails[disc_code] = st.text_input(
                    disc_cfg["label"],
                    value=emails.get(disc_code, ""),
                    placeholder=f"{disc_cfg['label'].lower().replace(' ','.')}@company.com",
                    key=f"email_{disc_code}",
                )
            idx += 1
        st.session_state["layer4_emails"] = emails

        cfg = _get_smtp_config()
        smtp_ready = bool(cfg["user"] and cfg["password"] and (cfg.get("from_email") or cfg["user"]))
        if smtp_ready:
            st.success(
                f"SMTP ready via {cfg.get('provider','SMTP')}: "
                f"{cfg.get('from_email') or cfg['user']} on {cfg['host']}:{cfg['port']}"
            )
        else:
            st.error(
                "SMTP2GO is not fully configured. Add SMTP2GO_HOST, SMTP2GO_PORT, "
                "SMTP2GO_USER, SMTP2GO_PASSWORD, and SMTP2GO_FROM to Streamlit secrets."
            )

        llm_key_available = bool(_get_deepseek_key())
        if llm_key_available and DeepSeekClient is not None:
            st.success("LLM drafting enabled: discipline routing and narrative Pre-RFI fields will be generated with DeepSeek.")
        else:
            st.warning("LLM drafting key not detected. K1netix will use deterministic fallback drafting.")

        st.markdown("---")
        st.markdown("### 4. Generate Pre-RFI Workbook")
        st.info(
            f"Will generate **{len(actionable)}** Pre-RFI form sheet(s). "
            "The workbook opens on the first filled form, then includes a register and evidence appendix."
        )

        if st.button("⚙️ Generate Pre-RFIs", type="primary"):
            prerfi_list = []
            prog = st.progress(0)
            status_box = st.empty()
            template_summary = st.session_state.get("layer4_template_summary", "")
            template_bytes = st.session_state.get("layer4_template_bytes")

            for i, cluster in enumerate(actionable):
                label = _s(cluster.get("cluster_label", "?"), 60)
                status_box.info(f"Drafting and routing Pre-RFI [{i+1}/{len(actionable)}]: {label}...")

                rfi_num = f"RFI-{datetime.now().strftime('%Y%m%d')}-{i+1:03d}"
                data = extract_prerfi_data(
                    cluster, rfi_num, project_name, "",
                    project_location=project_location,
                    project_id=project_id,
                    bim_model_version=bim_version,
                    clash_source_tool=clash_source,
                )
                data["prepared_by"] = prepared_by
                data = enhance_prerfi_with_llm(data, cluster, template_summary)
                data = apply_prerfi_routing(data, emails, cluster)
                prerfi_list.append({"data": data})
                prog.progress((i + 1) / max(len(actionable), 1))

            status_box.empty()
            st.session_state["layer4_prerfi_list"] = prerfi_list

            if prerfi_list:
                xlsx_bytes = build_xlsx_output(prerfi_list, template_bytes=template_bytes)
                st.session_state["layer4_xlsx_bytes"] = xlsx_bytes
                st.success(f"Generated {len(prerfi_list)} routed Pre-RFI form(s)")
                st.download_button(
                    "⬇️ Download Pre-RFI Workbook",
                    data=xlsx_bytes,
                    file_name=f"k1netix_prerfi_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

        prerfi_list = st.session_state.get("layer4_prerfi_list", [])
        template_bytes = st.session_state.get("layer4_template_bytes")
        if prerfi_list:
            # Keep routing in sync if the user changes a contact after generation.
            for p in prerfi_list:
                p["data"] = apply_prerfi_routing(p["data"], emails)
            st.session_state["layer4_xlsx_bytes"] = build_xlsx_output(prerfi_list, template_bytes=template_bytes)

            st.markdown("---")
            st.markdown("### 5. Preview, Routing & Send")
            preview_rows = []
            for p in prerfi_list:
                d = p["data"]
                preview_rows.append({
                    "RFI #": d.get("rfi_number", ""),
                    "Route": _discipline_label(d.get("_route_discipline", "")),
                    "Email": d.get("assigned_email", ""),
                    "Status": d.get("compliance_status", ""),
                    "Priority": d.get("priority_band", ""),
                    "Primary Concern": d.get("primary_concern", "")[:90],
                })
            st.dataframe(pd.DataFrame(preview_rows), hide_index=True, use_container_width=True)

            with st.expander("Full Pre-RFI content preview", expanded=False):
                for p in prerfi_list:
                    d = p["data"]
                    st.markdown(f"#### {d.get('rfi_number','')} — {d.get('cluster_label','')}")
                    for section_title, fields in PRE_RFI_SECTIONS:
                        st.markdown(f"**{section_title}**")
                        for field in fields:
                            label = FIELD_LABELS.get(field, field)
                            val = d.get(field, "")
                            if val not in (None, "", []):
                                st.markdown(f"- **{label}:** {val}")

            st.download_button(
                "⬇️ Download Updated Routed Workbook",
                data=st.session_state["layer4_xlsx_bytes"],
                file_name=f"k1netix_prerfi_routed_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

            disc_groups = _group_prerfis_by_route(prerfi_list, emails)
            # ── DEBUG: see what's going wrong with routing
            with st.expander("🔍 Routing debug"):
                st.write("**Emails dict (from Contacts tab):**")
                st.json({k: v for k, v in emails.items() if v})
                st.write("**Resolved route for each Pre-RFI:**")
                debug_rows = []
                for p in prerfi_list:
                    d = p["data"]
                    rc = _resolve_route_code(d)
                    debug_rows.append({
                        "RFI #":            d.get("rfi_number", ""),
                        "_route_discipline": d.get("_route_discipline", ""),
                        "_discipline":      d.get("_discipline", ""),
                        "discipline_label": d.get("discipline_label", ""),
                        "Resolved route_code": rc,
                        "Email found":      emails.get(rc, "❌ not in emails dict"),
                    })
                st.dataframe(pd.DataFrame(debug_rows), hide_index=True, use_container_width=True)
                
            route_rows = []
            for key, group in disc_groups.items():
                if key == "__unassigned__":
                    label, email_str = "Unassigned", "No email configured"
                else:
                    label = _discipline_label(key)
                    email_str = emails.get(key, "")
                route_rows.append({
                    "Discipline": label,
                    "Email": email_str,
                    "Pre-RFIs": len(group),
                    "RFI Numbers": ", ".join(p["data"].get("rfi_number", "") for p in group),
                })
            st.markdown("#### Routing Preview")
            st.dataframe(pd.DataFrame(route_rows), hide_index=True, use_container_width=True)

            if "__unassigned__" in disc_groups:
                st.warning(f"{len(disc_groups['__unassigned__'])} Pre-RFI(s) have no route email. Fill the matching contact above, then regenerate or update this page.")

            confirm = st.checkbox("I confirm the routing above is correct and I am ready to send", key="l4_send_confirm")
            can_send = confirm and smtp_ready and "__unassigned__" not in disc_groups
            if st.button("📤 Send Pre-RFIs to Specialists", type="primary", disabled=not can_send):
                send_log = []
                with st.spinner("Sending Pre-RFI emails..."):
                    for key, group in disc_groups.items():
                        if key == "__unassigned__":
                            continue
                        recipient = emails.get(key, "")
                        disc_label = _discipline_label(key)
                        subject = f"K1netix AI Pre-RFI for Review - {disc_label} - {datetime.now().strftime('%d %b %Y')}"
                        body = _email_body(group, disc_label)
                        group_bytes = build_xlsx_output(group, template_bytes=template_bytes)
                        r = send_email(
                            recipient, subject, body,
                            attachment_bytes=group_bytes,
                            attachment_name=f"K1netix_PreRFI_{disc_label.replace(' ', '_')}.xlsx",
                        )
                        entry = {
                            "Discipline": disc_label,
                            "Recipient": recipient,
                            "Pre-RFIs": len(group),
                            "Result": "Sent" if r["ok"] else f"Failed: {r.get('error','')}",
                            "Timestamp": datetime.now().strftime("%d/%m/%Y %H:%M"),
                        }
                        send_log.append(entry)
                        if r["ok"]:
                            st.success(f"Sent to {disc_label} specialist: {recipient}")
                        else:
                            st.error(f"{disc_label} email failed: {r.get('error','')}")

                st.session_state["layer4_records"].append({
                    "project": st.session_state.get("layer4_project_name", ""),
                    "total_rfis": len(prerfi_list),
                    "send_log": send_log,
                    "timestamp": datetime.now().isoformat(),
                    "prerfi_data": [p["data"] for p in prerfi_list],
                })

    with tabs[1]:
        st.markdown("### Dispatch Records")
        records = st.session_state.get("layer4_records", [])

        if not records:
            st.info("No records yet. Send Pre-RFIs to create records.")
        else:
            st.markdown(f"**{len(records)} dispatch batch(es)** on record.")
            for i, rec in enumerate(reversed(records)):
                ts = rec.get("timestamp", "")[:19].replace("T", " ")
                proj = rec.get("project", "")
                total = rec.get("total_rfis", 0)
                with st.expander(f"Batch {len(records)-i} · {proj} · {ts} · {total} Pre-RFIs"):
                    log = rec.get("send_log", [])
                    if log:
                        st.dataframe(pd.DataFrame(log), hide_index=True, use_container_width=True)
                    data_list = rec.get("prerfi_data", [])
                    if data_list:
                        skip = {"_cluster_id", "_generated_at", "_llm_drafting"}
                        flat = [{k: v for k, v in d.items() if k not in skip} for d in data_list]
                        csv = pd.DataFrame(flat).to_csv(index=False)
                        st.download_button(
                            f"⬇️ Download Batch {len(records)-i} CSV",
                            data=csv,
                            file_name=f"k1netix_batch_{len(records)-i}.csv",
                            mime="text/csv",
                            key=f"dl_{i}",
                        )

            all_data = []
            for rec in records:
                all_data.extend(rec.get("prerfi_data", []))
            if all_data:
                st.markdown("---")
                st.download_button(
                    "⬇️ Export All Records (CSV)",
                    data=pd.DataFrame(all_data).to_csv(index=False),
                    file_name="k1netix_all_prerfi.csv",
                    mime="text/csv",
                )
